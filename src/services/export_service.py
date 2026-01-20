"""
Export service for generating CSV and Excel files.

References:
- FR-072: Data export capability
"""

import csv
import io
from datetime import date
from typing import Optional

from sqlalchemy.orm import Session

from src.models.case import CaseType, CaseStatus
from src.services.report_service import ReportService


class ExportService:
    """
    Service for exporting data to various formats.
    """

    def __init__(self, db: Session):
        """
        Initialize export service.

        Args:
            db: Database session
        """
        self.db = db
        self.report_service = ReportService(db)

    def export_cases_csv(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        case_type: Optional[CaseType] = None,
        status: Optional[CaseStatus] = None
    ) -> bytes:
        """
        Export cases to CSV format.

        Args:
            start_date: Start of period filter
            end_date: End of period filter
            case_type: Optional case type filter
            status: Optional status filter

        Returns:
            CSV file content as bytes
        """
        cases = self.report_service.get_cases_for_export(
            start_date=start_date,
            end_date=end_date,
            case_type=case_type,
            status=status
        )

        if not cases:
            return b"No data to export"

        # Get headers from first case
        headers = list(cases[0].keys())

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        writer.writerows(cases)

        return output.getvalue().encode('utf-8')

    def export_cases_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        case_type: Optional[CaseType] = None,
        status: Optional[CaseStatus] = None
    ) -> bytes:
        """
        Export cases to Excel format.

        Args:
            start_date: Start of period filter
            end_date: End of period filter
            case_type: Optional case type filter
            status: Optional status filter

        Returns:
            Excel file content as bytes
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            # Fallback to CSV if openpyxl not available
            return self.export_cases_csv(
                start_date=start_date,
                end_date=end_date,
                case_type=case_type,
                status=status
            )

        cases = self.report_service.get_cases_for_export(
            start_date=start_date,
            end_date=end_date,
            case_type=case_type,
            status=status
        )

        if not cases:
            return b"No data to export"

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Case Export"

        # Headers
        headers = list(cases[0].keys())
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font

        # Data rows
        for row_idx, case_data in enumerate(cases, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=case_data.get(header, ""))

        # Auto-adjust column widths
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def get_export_filename(
        self,
        export_format: str,
        prefix: str = "case_export"
    ) -> str:
        """
        Generate export filename with timestamp.

        Args:
            export_format: File format (csv, xlsx)
            prefix: Filename prefix

        Returns:
            Generated filename
        """
        from datetime import datetime
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.{export_format}"
